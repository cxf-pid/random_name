
import json
import random
from openpyxl import load_workbook
import os
import sys
from tkinter import Tk, Button, Label, filedialog, BOTH, BOTTOM, TOP, messagebox

cache_file = "students_cache.json"

def load_from_cache():
    if os.path.exists(cache_file):
        with open(cache_file, 'r') as file:
            global students
            students = json.load(file)
        display_message("从缓存加载数据成功！")
    else:
        display_message("缓存文件不存在，需要加载新的Excel文件。")

def save_to_cache():
    with open(cache_file, 'w') as file:
        json.dump(students, file)
    display_message("数据已缓存！")

def random_select():
    global i
    if students:
        available_students = [s for s in students if s[0] not in selected_students]
        if available_students:
            selected_student = random.choice(available_students)
            selected_students.append(selected_student[0])
            i += 1
            display_message(f"第{i}人，姓名：{selected_student[1]}，学号：{selected_student[0]}")
        else:
            display_message("所有学生都已点过名。")
    else:
        display_message("没有学生数据，请先导入Excel文件。")

selected_students = []
i = 0

def load_excel():
    global students
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        students = [list(row) for row in sheet.iter_rows(min_row=1, values_only=True)]
        save_to_cache()
        display_message("文件加载成功！")

def display_message(message):
    message_label.config(text=message, font=('Helvetica',36))

def JieShu():
    messagebox.showwarning(title='警告', message='刚才你点击了关闭按钮')
    sys.exit(0)
    root.destroy()

root = Tk()
root.title("随机点名程序")
root.geometry("1000x300+300+300")

load_button = Button(root, text="导入数据", command=load_excel)
load_button.pack(side=TOP, padx=0, pady=0)

message_label = Label(root, text="", font=('Helvetica', 36), bg='black', fg='white')
message_label.pack(expand=True, fill=BOTH)

select_button = Button(root, text="随机点名", font=('Helvetica', 20), command=random_select)
select_button.pack(side=BOTTOM, pady=20)

load_from_cache()
root.title("随机点名")
root.protocol("WM_DELETE_WINDOW", JieShu)
root.mainloop()
